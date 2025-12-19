#!/usr/bin/env python3

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

# Create the workbook
wb = Workbook()
wb.remove(wb.active)

# Define colors
colors = {
    'primary': '2E86AB',
    'secondary': 'A23B72', 
    'accent': 'F18F01',
    'success': 'C73E1D',
    'neutral': '7D8491'
}

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 1. Application Tracker Sheet
ws1 = wb.create_sheet('Application Tracker')

# Create sample data
sample_data = {
    'Company': ['GitLab', 'Basecamp', 'Zapier', 'Stripe', 'Monzo', 'Wise', 'Discord', 'Automattic', 'Shopify', 'Revolut'],
    'Role Title': ['Junior Backend Developer', 'Junior Programmer', 'Junior Engineer', 'Junior Backend Engineer', 'Junior Backend Engineer', 'Junior Software Engineer', 'Junior Backend Engineer', 'Junior Developer', 'Junior Developer', 'Junior Developer'],
    'Location': ['Remote'] * 10,
    'Hourly Rate ($)': [45, 50, 48, 55, 42, 40, 52, 46, 44, 43],
    'Applied Date': pd.date_range('2024-12-18', periods=10),
    'Application Source': ['Company Website', 'LinkedIn', 'Wellfound', 'Referral', 'LinkedIn', 'Company Website', 'LinkedIn', 'Company Website', 'LinkedIn', 'Wellfound'],
    'Status': ['Applied', 'Applied', 'Screening', 'Applied', 'Applied', 'Applied', 'Applied', 'Applied', 'Applied', 'Applied'],
    'Last Update': pd.date_range('2024-12-18', periods=10),
    'Contact Person': ['Tech Recruiter', 'HR Manager', 'Engineering Manager', 'Employee Referral', 'Recruitment Team', 'Tech Recruiter', 'Engineering Manager', 'HR Team', 'Recruiter', 'Talent Team'],
    'Contact Email/LinkedIn': ['recruiter@gitlab.com', 'hr@basecamp.com', 'em@zapier.com', 'friend@stripe.com', 'jobs@monzo.com', 'careers@wise.com', 'jobs@discord.com', 'work@automattic.com', 'careers@shopify.com', 'jobs@revolut.com'],
    'Interview Date': [''] * 10,
    'Interview Type': [''] * 10,
    'Follow-up Date': pd.date_range('2024-12-25', periods=10),
    'Notes': ['Remote-first, great culture', 'Calm company', 'Automation platform', 'Payment processing', 'Digital bank', 'Fintech', 'Communication platform', 'WordPress', 'E-commerce', 'Digital bank'],
    'Priority Level': ['High'] * 3 + ['Medium'] * 7,
    'Company Size': ['Large', 'Small', 'Medium', 'Large', 'Medium', 'Large', 'Large', 'Large', 'Large', 'Large'],
    'Tech Stack': ['Ruby, Go, React', 'Ruby on Rails', 'Python, React', 'Ruby, Go', 'Go, React', 'Java, React', 'Elixir, Python', 'PHP, JavaScript', 'Ruby, React', 'Java, React'],
    'Application Method': ['Direct Application', 'Easy Apply', 'Direct Application', 'Employee Referral', 'Easy Apply', 'Direct Application', 'Easy Apply', 'Direct Application', 'Easy Apply', 'Direct Application']
}

df = pd.DataFrame(sample_data)

# Add headers and data
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        if r_idx == 1:  # Header row
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Set column widths
for col in range(1, 19):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

# 2. Dashboard Sheet
ws2 = wb.create_sheet('Dashboard')

# Dashboard title
title_cell = ws2['A1']
title_cell.value = 'Job Search Analytics Dashboard'
title_cell.font = Font(bold=True, size=16, color='FFFFFF')
title_cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type='solid')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws2.merge_cells('A1:D1')
ws2.row_dimensions[1].height = 30

# Key metrics
metrics = [
    ('Total Applications', len(df)),
    ('Applications This Week', 10),
    ('Interviews Scheduled', 2),
    ('Offers Received', 0),
    ('High Priority Applications', 3),
    ('Average Hourly Rate', f'${df["Hourly Rate ($)"].mean():.0f}'),
    ('Applications Today', 5),
    ('Success Rate (%)', '0%')
]

for i, (label, value) in enumerate(metrics):
    row = 3 + (i // 4) * 3
    col = (i % 4) + 1
    
    # Label
    label_cell = ws2.cell(row=row, column=col)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=10)
    label_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Value
    value_cell = ws2.cell(row=row+1, column=col)
    value_cell.value = value
    value_cell.font = Font(bold=True, size=14, color='FFFFFF')
    value_cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type='solid')
    value_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws2.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)
    ws2.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col)
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

# 3. Target Companies Sheet
ws3 = wb.create_sheet('Target Companies')

target_companies_data = {
    'Company Name': ['GitLab', 'Basecamp', 'Zapier', 'Stripe', 'Discord', 'Monzo', 'Wise', 'Automattic', 'Shopify', 'Revolut', 'ThoughtWorks', 'Sparta Global', 'ECA International', 'SmartChoice International', 'The Dev Hub'],
    'Role Title': ['Junior Backend Developer', 'Junior Programmer', 'Junior Engineer', 'Junior Backend Engineer', 'Junior Backend Engineer', 'Junior Backend Engineer', 'Junior Software Engineer', 'Junior Developer', 'Junior Developer', 'Junior Developer', 'Junior Developer', 'Junior Java Developer', 'Front-End Developer', 'Junior Java Developer', 'Web Developer'],
    'Website/Careers Page': ['https://about.gitlab.com/jobs/', 'https://basecamp.com/about/jobs', 'https://zapier.com/jobs/', 'https://stripe.com/jobs', 'https://discord.com/jobs', 'https://monzo.com/careers/', 'https://wise.jobs/', 'https://automattic.com/work-with-us/', 'https://shopify.com/careers', 'https://jobs.lever.co/revolut', 'https://thoughtworks.com/careers', 'https://spartaglobal.com/careers', 'https://eca-international.com/careers', 'https://smartchoice.global/careers', 'https://thedevhub.io/careers'],
    'Company Size': ['Large', 'Small', 'Medium', 'Large', 'Large', 'Medium', 'Large', 'Large', 'Large', 'Large', 'Large', 'Medium', 'Medium', 'Medium', 'Small'],
    'Industry': ['DevOps', 'Project Management', 'Automation', 'Fintech', 'Communication', 'Fintech', 'Fintech', 'Web Publishing', 'E-commerce', 'Fintech', 'Consulting', 'Training', 'Business Intelligence', 'Technology', 'Web Development'],
    'Remote Policy': ['Fully Remote', 'Fully Remote', 'Fully Remote', 'Remote Friendly', 'Remote Friendly', 'Remote Friendly', 'Remote Friendly', 'Fully Remote', 'Remote Friendly', 'Remote Friendly', 'Remote Friendly', 'Hybrid', 'Remote Friendly', 'Hybrid', 'Fully Remote'],
    'Tech Stack': ['Ruby, Go, React', 'Ruby on Rails', 'Python, React', 'Ruby, Go', 'Elixir, Python', 'Go, React', 'Java, React', 'PHP, JavaScript', 'Ruby, React', 'Java, React', 'Multiple', 'Java', 'JavaScript', 'Java', 'Multiple'],
    'Application Status': ['Not Applied'] * 15,
    'Priority': ['High'] * 7 + ['Medium'] * 8,
    'Notes': ['Remote-first, great culture', 'Calm company philosophy', 'Automation platform leader', 'Payment processing leader', 'Communication platform', 'Digital bank', 'International payments', 'WordPress', 'E-commerce platform', 'Digital bank', 'Global consulting', 'Junior training program', '6-month contract', 'Technology academy', 'Developer community']
}

target_df = pd.DataFrame(target_companies_data)

for r_idx, row in enumerate(dataframe_to_rows(target_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws3.cell(row=r_idx, column=c_idx)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        if r_idx == 1:
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color=colors['secondary'], end_color=colors['secondary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

for col in range(1, 11):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

# 4. Interview Prep Sheet
ws4 = wb.create_sheet('Interview Prep')

interview_data = {
    'Company': ['GitLab', 'Zapier', 'Stripe'],
    'Role': ['Junior Backend Developer', 'Junior Engineer', 'Junior Backend Engineer'],
    'Interview Date': ['2024-12-23', '2024-12-24', '2024-12-26'],
    'Interview Type': ['Phone Screen', 'Technical Interview', 'Final Round'],
    'Interviewer': ['Engineering Manager', 'Tech Lead', 'CTO'],
    'Questions to Ask': ['Team structure? Mentorship?', 'Tech stack? Learning opportunities?', 'Company growth? Career path?'],
    'Key Points to Mention': ['PhD advantage, remote experience', 'Python experience, automation interest', 'Payment experience, security focus'],
    'Research Notes': ['GitLab remote culture', 'Zapier automation platform', 'Stripe payment processing'],
    'Technical Prep': ['Ruby basics, system design', 'Python coding challenge', 'System design, security'],
    'Follow-up Actions': ['Send thank you email', 'Prepare coding examples', 'Review Stripe API docs']
}

interview_df = pd.DataFrame(interview_data)

for r_idx, row in enumerate(dataframe_to_rows(interview_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws4.cell(row=r_idx, column=c_idx)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color=colors['accent'], end_color=colors['accent'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if r_idx > 1:
            ws4.row_dimensions[r_idx].height = 60

for col in range(1, 11):
    ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

# 5. Rate Negotiation Sheet
ws5 = wb.create_sheet('Rate Negotiation')

negotiation_data = {
    'Company': ['GitLab', 'Zapier', 'Stripe'],
    'Role': ['Junior Backend Developer', 'Junior Engineer', 'Junior Backend Engineer'],
    'Initial Offer ($)': [38, 42, 45],
    'Target Rate ($)': [45, 48, 55],
    'Minimum Acceptable ($)': [40, 42, 45],
    'Negotiation Points': ['PhD + experience', 'Remote experience', 'Security background'],
    'Counter Offer Strategy': ['Emphasize unique value', 'Trial period option', 'Performance review clause'],
    'Final Rate ($)': ['', '', ''],
    'Decision': ['', '', ''],
    'Notes': ['Great culture', 'Automation leader', 'High growth']
}

negotiation_df = pd.DataFrame(negotiation_data)

for r_idx, row in enumerate(dataframe_to_rows(negotiation_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws5.cell(row=r_idx, column=c_idx)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        if r_idx == 1:
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

for col in range(1, 11):
    ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

# 6. Daily Log Sheet
ws6 = wb.create_sheet('Daily Log')

log_data = {
    'Date': pd.date_range('2024-12-18', periods=14),
    'Applications Sent': [20, 22, 18, 25, 23, 21, 19, 24, 26, 20, 22, 25, 23, 21],
    'LinkedIn Connections': [15, 18, 12, 20, 17, 16, 14, 19, 21, 15, 18, 20, 17, 16],
    'Recruiter Messages': [3, 4, 2, 5, 4, 3, 3, 4, 5, 3, 4, 4, 4, 3],
    'Interviews Scheduled': [0, 0, 1, 0, 0, 1, 0, 0, 1, 0, 0, 1, 0, 1],
    'Hours Spent': [8, 9, 7, 10, 9, 8, 7, 9, 10, 8, 9, 10, 9, 8],
    'Key Activities': ['Applied to 20 companies, networked on LinkedIn'] * 14,
    'Tomorrow\'s Goals': ['Apply to 20+ companies'] * 14,
    'Notes': [f'Day {i} of job search' for i in range(1, 15)]
}

log_df = pd.DataFrame(log_data)

for r_idx, row in enumerate(dataframe_to_rows(log_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws6.cell(row=r_idx, column=c_idx)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        if r_idx == 1:
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color=colors['neutral'], end_color=colors['neutral'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

for col in range(1, 10):
    ws6.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

# Save the workbook
wb.save('Hamman_Job_Search_Tracker.xlsx')

print('✅ Excel workbook created successfully!')
print('📊 File saved as: Hamman_Job_Search_Tracker.xlsx')
print('📋 Sheets included:')
print('   1. Application Tracker - Track all job applications')
print('   2. Dashboard - Analytics and metrics')
print('   3. Target Companies - Priority company list')
print('   4. Interview Prep - Interview preparation notes')
print('   5. Rate Negotiation - Salary negotiation tracking')
print('   6. Daily Log - Daily activity tracking')
